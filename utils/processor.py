import pandas as pd
import warnings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from unidecode import unidecode
import PyPDF2
import os
import logging
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

class FileValidationError(Exception):
    pass

class ProcessingError(Exception):
    pass

def validate_files(arquivo_xlsx, arquivo_pdf, parametros_xlsx):
    """Valida se os arquivos estão no formato correto"""
    errors = []
    
    # Validar arquivo Excel principal
    try:
        df_test = pd.read_excel(arquivo_xlsx, engine='openpyxl')
        required_columns = ['DIA', 'LOCADOR', 'IMÓVEL', 'REFERÊNCIA', 'DESCRIÇÃO', 
                          'COMPLEMENTO HISTÓRICO', 'LOCATÁRIO', 'VALOR RECEBIDO', 'VALOR PAGO', 
                          'ENDEREÇO DO IMÓVEL', 'TIPO DE LANÇAMENTO']
        
        missing_columns = [col for col in required_columns if col not in df_test.columns]
        if missing_columns:
            errors.append(f"Arquivo Excel principal: Colunas obrigatórias ausentes: {', '.join(missing_columns)}")
    except Exception as e:
        errors.append(f"Erro ao ler arquivo Excel principal: {str(e)}")
    
    # Validar arquivo de parâmetros
    try:
        # Verificar se as abas existem
        excel_file = pd.ExcelFile(parametros_xlsx)
        required_sheets = ['Listas', 'Parametrização']
        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_file.sheet_names]
        if missing_sheets:
            errors.append(f"Arquivo de parâmetros: Abas obrigatórias ausentes: {', '.join(missing_sheets)}")
        
        # Verificar colunas da aba Parametrização
        if 'Parametrização' in excel_file.sheet_names:
            df_params = pd.read_excel(parametros_xlsx, sheet_name="Parametrização", skiprows=[0], engine='openpyxl')
            required_param_columns = ['Locais (Do CSV)', 'Locais (Para Quickbooks)', 
                                    'Locatário (De PDF)', 'Locatário (Para Quickbooks)',
                                    'Descrição (do CSV)', 'Centro de Custo', 'Categoria de Despesa']
            missing_param_columns = [col for col in required_param_columns if col not in df_params.columns]
            if missing_param_columns:
                errors.append(f"Arquivo de parâmetros (aba Parametrização): Colunas obrigatórias ausentes: {', '.join(missing_param_columns)}")
    except Exception as e:
        errors.append(f"Erro ao ler arquivo de parâmetros: {str(e)}")
    
    # Validar PDF
    try:
        with open(arquivo_pdf, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            if len(reader.pages) == 0:
                errors.append("Arquivo PDF está vazio ou corrompido")
    except Exception as e:
        errors.append(f"Erro ao ler arquivo PDF: {str(e)}")
    
    if errors:
        raise FileValidationError("; ".join(errors))
    
    return True

def split_pdf_and_rotate(input_pdf, output_pdf):
    """Função para dividir uma página do PDF horizontalmente e deixar as metades em pé"""
    with open(input_pdf, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        writer = PyPDF2.PdfWriter()

        for page in reader.pages:
            # Obter dimensões da página
            width = page.mediabox.width
            height = page.mediabox.height

            # Criar a página superior e rotacionar
            top_page = PyPDF2.PageObject.create_blank_page(width=height / 2, height=width)
            top_page.merge_page(page)
            top_page.mediabox.lower_left = (0, height / 2)
            top_page.mediabox.upper_right = (width, height)
            top_page.rotate(90)
            writer.add_page(top_page)

            # Criar a página inferior e rotacionar
            bottom_page = PyPDF2.PageObject.create_blank_page(width=height / 2, height=width)
            bottom_page.merge_page(page)
            bottom_page.mediabox.lower_left = (0, 0)
            bottom_page.mediabox.upper_right = (width, height / 2)
            bottom_page.rotate(90)
            writer.add_page(bottom_page)

        # Salvar o novo PDF com as páginas ajustadas
        with open(output_pdf, "wb") as output_file:
            writer.write(output_file)

def process_files(arquivo_xlsx, arquivo_pdf, parametros_xlsx, nome_cliente, mes_ano, transferencia, pagamentos, saldo_inicial, saldo_final, diario_n, output_dir):
    """Processa os arquivos conforme a lógica do app1.py"""
    
    try:
        # Validar arquivos primeiro
        validate_files(arquivo_xlsx, arquivo_pdf, parametros_xlsx)
        
        # Carregar dados
        df_list = pd.read_excel(parametros_xlsx, sheet_name="Listas", keep_default_na=False, engine='openpyxl')
        df_base = pd.read_excel(arquivo_xlsx, converters={'str':['DIA', 'LOCADOR', 'IMÓVEL', 'LOCADOR','REFERÊNCIA','DESCRIÇÃO', 'COMPLEMENTO HISTÓRICO','LOCATÁRIO']}, engine='openpyxl', keep_default_na=False, sheet_name=[0])[0]
        initial_columns = df_base.columns

        # Limpar dados
        for i in range(len(df_base)):
            for c in range(len(df_base.columns)):
                df_base.iloc[i, c] = str(df_base.iloc[i, c]).strip()

        df_base['Data'] = ['0' + str(d) + '/' + mes_ano if len(str(d)) == 1 else str(d) + '/' +  mes_ano for d in df_base['DIA']]
        df_base['Complemento_Descricao'] = [" ".join([des.strip(), ch.strip(), ref.strip()]) for des, ch, ref in df_base[['DESCRIÇÃO', 'COMPLEMENTO HISTÓRICO', 'REFERÊNCIA']].values]
        df_base['VALOR RECEBIDO'] = ["-" + str(v).strip() if  ("dev." in desc.strip().lower() or "desc. aluguel" in desc.strip().lower()) 
                                    and str(desc).strip().lower() not in ['', 'nan'] else str(v).strip() for v, desc in df_base[['VALOR RECEBIDO', 'DESCRIÇÃO']].values]
        
        df_base['VALOR PAGO'] = [str(v).strip() for v in df_base['VALOR PAGO']]

        # Ajustar Locais
        df_params = pd.read_excel(parametros_xlsx, sheet_name="Parametrização", keep_default_na=False, skiprows=[0], engine='openpyxl')

        list_locais_csv = df_params['Locais (Do CSV)'].to_list()
        list_locais_csv = [ unidecode(str(v.split(' ', 1)[1].strip()).upper()) if v.split(' ', 1)[0].replace('-','').strip().isnumeric() else unidecode(v.strip().upper()) for v in list_locais_csv ]

        list_locais_quickbooks = df_params['Locais (Para Quickbooks)'].to_list()
        list_locais_quickbooks = [v.strip() for v in list_locais_quickbooks]
        df_base['ENDEREÇO DO IMÓVEL'] = [ list_locais_quickbooks[list_locais_csv.index(unidecode(v.strip().upper()))] if str(v).strip().upper() not in ['', 'NAN', 'NONE'] and unidecode(str(v).strip().upper()) in list_locais_csv else v.strip() for v in df_base['ENDEREÇO DO IMÓVEL'] ]

        # Processar PDF
        output_pdf = os.path.join(output_dir, "saida.pdf")
        split_pdf_and_rotate(arquivo_pdf, output_pdf)

        columns = ['locatario','code_imovel']
        df_pdf = pd.DataFrame(columns=columns)

        with open(output_pdf, "rb") as file:
            reader = PyPDF2.PdfReader(file)

            # Iterar pelas páginas e extrair o texto
            for page_num, page in enumerate(reader.pages, start=1):
                text = page.extract_text()
                v=0
                for row in text.splitlines():
                    row_split = row.split(' ')
                    code = row_split[0].strip()
                    
                    if code.replace('-', '').isnumeric() and '-' in code:
                        row_split_ = row.split(' - ')
                        
                        if len(row_split_) >= 2 and not '(cont inuação)' in row_split_[-1].strip() :
                            v += 1
                            
                            if v >= 1:
                                locatario = row_split_[-1].strip()
                                locatario = locatario[:3] + locatario[4:]  if locatario[3] == ' ' else locatario
                                locatario = " ".join([v for v in locatario.split(' ')])
                                code = int(code.replace('-',''))
                                
                                df_pdf.loc[len(df_pdf), df_pdf.columns] = [locatario, code]
        
        df_pdf = df_pdf.drop_duplicates()
        df_pdf['code_imovel'] = df_pdf['code_imovel'].apply(str)
        df_pdf['Auxiliar_Vago'] = ["sim" if v.lower().strip() == "vago" else "não" for v in df_pdf.locatario]
        df_pdf.sort_values(by='locatario')

        # Ajustar Locatarios
        list_locatarios_pdf = df_params['Locatário (De PDF)'].to_list()
        list_locatarios_pdf = [v.strip().upper() for v in list_locatarios_pdf ]

        list_locatarios_quickbooks = df_params['Locatário (Para Quickbooks)'].to_list()
        list_locatarios_quickbooks = [v.strip() for v in list_locatarios_quickbooks]
        list_locatarios_quickbooks_upper = [v.strip().upper() for v in list_locatarios_quickbooks]

        df_pdf['locatario'] = [str(list_locatarios_quickbooks[list_locatarios_pdf.index(str(v).strip().upper())]) if str(v).strip().upper() in list_locatarios_pdf and 
                                str(list_locatarios_quickbooks_upper[list_locatarios_pdf.index(str(v).strip().upper())]) not in ['', 'None', 'NAN'] 
                                else v.strip().upper() for v in df_pdf['locatario'].to_list() ]

        # Mesclar Dados do PDF
        df_base['Auxiliar_Vago'] = ["sim" if v.strip() == "" else "não" for v in df_base.LOCATÁRIO]
        df_base = df_base.merge(df_pdf, how="left", right_on=["code_imovel", "Auxiliar_Vago"], left_on=["IMÓVEL","Auxiliar_Vago"])
        
        # Centro de Custo
        idx_desc = list(df_base.columns).index('DESCRIÇÃO')
        list_history = df_params['Descrição (do CSV)']
        list_ctc = df_params['Centro de Custo']
        list_categoria = df_params['Categoria de Despesa']
        df_base['Categoria_Auxiliar'] = [None for _ in df_base.index]
        df_base['Centro de Custo'] = [None for _ in df_base.index]

        for i in range(len(df_base)):
            value_base = unidecode(df_base.iloc[i, idx_desc]).strip().lower()
            
            for k,value in enumerate(list_history):
                value_params = unidecode(value).strip().lower()
                if value_params == value_base:
                    df_base.iloc[i, -1] = list_ctc[k]
                    df_base.iloc[i, -2] = list_categoria[k]
                    
                # Ajuste manual dos diversos
                elif value_base.startswith('diversos'):
                    df_base.iloc[i, -1] = "Manutenção"
                    df_base.iloc[i, -2] = "Custos com imóveis próprios:Manutenção e Condomínio"

        # Processar receitas
        df_receitas_aluguel = df_base[df_base['TIPO DE LANÇAMENTO'].str.strip().str.lower() == "recibos de venda"]
        df_receitas_aluguel['Conta'] = 'Caixa e equivalentes-caixa:Adibras'
        df_receitas_aluguel['VALOR PAGO'] = [float(v.strip().replace(',', '.')) if v.strip() != '' else 0 for v in df_receitas_aluguel['VALOR PAGO'] ]
        df_receitas_aluguel['VALOR RECEBIDO'] = [float(v.strip().replace(',', '.')) if v.strip() != '' else 0 for v in df_receitas_aluguel['VALOR RECEBIDO'] ]

        df_despesas_aluguel = df_receitas_aluguel[df_receitas_aluguel.DESCRIÇÃO.str.strip().str.lower() == "desc. aluguel"]
        df_receitas_aluguel['Auxiliar'] = [1 if v.strip().lower() in (["aluguel", "desc. aluguel"]) else 0 for v in df_receitas_aluguel['DESCRIÇÃO']]
        df_receitas_aluguel = df_receitas_aluguel[df_receitas_aluguel.DESCRIÇÃO.str.strip().str.lower() != "desc. aluguel"]

        df_despesas_aluguel['VALOR RECEBIDO'] = [v*-1 if v < 0 else v for v in df_despesas_aluguel['VALOR RECEBIDO']]
        df_despesas_aluguel = df_despesas_aluguel.groupby(["ENDEREÇO DO IMÓVEL", "locatario"], as_index=False).agg({'VALOR RECEBIDO': 'sum'})
        df_despesas_aluguel["Auxiliar"] = 1

        df_receitas_aluguel = df_receitas_aluguel.merge(df_despesas_aluguel, how="left", on=["ENDEREÇO DO IMÓVEL", "locatario", "Auxiliar"])

        list_columns = [ "locatario", "Data", "Conta", "DESCRIÇÃO", "Complemento_Descricao", "VALOR RECEBIDO_x", 'VALOR RECEBIDO_y', "ENDEREÇO DO IMÓVEL" ]
        list_new_columns = ["Cliente", "Data do recibo de venda", "Conta", "Produto/Serviço", "Descrição", "Valor", "Desconto", "Local"]

        df_receitas_aluguel = df_receitas_aluguel[list_columns]
        df_receitas_aluguel.columns = list_new_columns

        df_receitas_aluguel['Desconto'] = df_receitas_aluguel['Desconto'].fillna(0)
        df_receitas_aluguel['Valor'] = df_receitas_aluguel['Valor'].fillna(0)
        
        df_receitas_aluguel['Saldo'] = [x - y if y > 0 else x + y for x, y in zip( df_receitas_aluguel['Valor'], df_receitas_aluguel['Desconto'] )]

        df_receitas_aluguel['Desconto'] = [float(str(v).replace('-', '')) if v != 0 else None for v in df_receitas_aluguel['Desconto']]
        df_receitas_aluguel['Valor'] = [float(v) if v != 0 else None for v in df_receitas_aluguel['Valor']]

        df_receitas_aluguel['Sales Receipt No'] = None

        list_new_columns = ["Sales Receipt No", "Cliente", "Data do recibo de venda", "Conta", "Produto/Serviço", "Descrição", "Valor", "Desconto", "Saldo","Local"]
        df_receitas_aluguel = df_receitas_aluguel[list_new_columns]
        df_receitas_aluguel["Produto/Serviço"] = "Aluguel"

        # Processar despesas
        df_despesas = df_base[df_base['TIPO DE LANÇAMENTO'].str.strip().str.lower() == "despesas"]

        df_despesas['Beneficiário'] = 'Adibras Administradora Brasileira de bens LTDA'
        df_despesas['Conta'] = 'Caixa e equivalentes-caixa:Adibras'
        df_despesas['Auxiliar_Valor'] = [rec.replace('-', '') if "-" in rec else pago for rec, pago in df_despesas[['VALOR RECEBIDO', 'VALOR PAGO']].values]
        df_despesas["Auxiliar_Valor"] = [float(str(v).replace(',', '.').strip()) if v != '' else None for v in df_despesas["Auxiliar_Valor"]]
        
        list_columns = [ "Conta", "Beneficiário", "Data", "Categoria_Auxiliar", "Complemento_Descricao",  "Auxiliar_Valor", "ENDEREÇO DO IMÓVEL", "Centro de Custo", "locatario"]
        list_new_columns = ["Conta", "Beneficiário", "Data do pagamento", "Categoria", "Descrição", "Valor", "Local", "Centro de Custo", "locatario"]

        df_despesas = df_despesas[list_columns]
        df_despesas.columns = list_new_columns
        df_despesas = df_despesas[df_despesas['Valor'].apply(str) != 'nan']

        # Processar lançamentos contábeis
        df_contabil = df_base[df_base['TIPO DE LANÇAMENTO'].str.strip().str.lower() == "lançamentos contábeis"]

        diario_n = diario_n - 1
        date_increment = pd.to_datetime('01/01/2023', format='%d/%m/%Y')
        list_diario = []

        for d in sorted(df_contabil['Data']):
            try: 
                date_ = pd.to_datetime(d, format='%d/%m/%Y')
                if date_ > date_increment:
                    diario_n += 1
                    list_diario.append(diario_n)
                    date_increment = date_
                else: list_diario.append(diario_n)
            except: 
                list_diario.append(diario_n)
        
        df_contabil['Diário nº'] = list_diario
        df_contabil['Conta'] = "Outros Passivos:Adibras - Transitória de passivos"

        list_columns = [ "Diário nº", "Data", "Conta", "VALOR PAGO", "VALOR RECEBIDO", "Complemento_Descricao", "locatario", "ENDEREÇO DO IMÓVEL", "Centro de Custo"]
        list_new_columns = ["Diário nº", "Data do lançamento", "Conta","Débito", "Crédito", "Descrição", "Nome", "Local", "Centro de Custo"]

        df_contabil["VALOR PAGO"] = [float(str(v).replace(',', '.').strip()) if v != '' else None for v in df_contabil["VALOR PAGO"]]
        df_contabil["VALOR RECEBIDO"] = [float(str(v).replace(',', '.').strip()) if v != '' else None for v in df_contabil["VALOR RECEBIDO"]]
        
        df_contabil = df_contabil[list_columns]
        df_contabil.columns = list_new_columns
        df_contabil = df_contabil.reset_index(drop=True)
        
        # Lançamentos Contábeis 2
        df_contabil2 = df_contabil.copy()
        df_contabil2['Crédito'] = [v for v in df_contabil['Débito']]
        df_contabil2['Débito'] = [v for v in df_contabil['Crédito']]
        df_contabil2['Conta'] = ["Outros Passivos:Adibras - Transitória de passivos" if c.strip() == "Adibras - Transitória de ativos" else "Adibras - Transitória de ativos" for c in df_contabil2['Conta']]
        df_contabil2 = pd.concat([df_contabil, df_contabil2], ignore_index=True).reset_index(drop=True)
        df_contabil2 = df_contabil2.sort_values(by=['Diário nº', "Data do lançamento"])

        # Processar faltantes
        df_faltantes = df_base[(df_base['TIPO DE LANÇAMENTO'].apply(str).str.strip() == 'Lançamentos manuais')]
        df_faltantes['VALOR RECEBIDO'] = [float(v.replace(',', '.')) if str(v).replace(',', '').replace('.', '').replace('-', '').isnumeric() else v for v in df_faltantes['VALOR RECEBIDO']]
        df_faltantes['VALOR PAGO'] = [float(v.replace(',', '.')) if str(v).replace(',', '').replace('.', '').replace('-', '').isnumeric() else v for v in df_faltantes['VALOR PAGO']]
        df_faltantes = df_faltantes[initial_columns]
        df_despesas.drop(columns=['locatario'], inplace=True)

        # Salvar arquivo
        def save_excel(path, list_dfs, list_name_sheets):
            # Copiar arquivo de parâmetros como base
            import shutil
            shutil.copy2(parametros_xlsx, path)

            for df, sheet_name in zip(list_dfs, list_name_sheets):
                wb = load_workbook(path)
                
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                
                wb.create_sheet(sheet_name)
                sheet = wb[sheet_name]
                
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)

                for column_cells in sheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 7
                        
                wb.save(path)
                wb.close()

        output_path = os.path.join(output_dir, f"Saida_Matriz_{nome_cliente.replace(' ', '_')}_{mes_ano.replace('/','_')}.xlsx")

        list_dfs = [df_receitas_aluguel, df_despesas, df_contabil2, df_faltantes]
        list_name_sheets = ['Recibos de Venda', 'Despesas', 'Lançamentos Contábeis', 'Lançamentos manuais']

        save_excel(output_path, list_dfs, list_name_sheets)

        # Criar saldos
        wb = load_workbook(output_path)
        ws = wb['Conferência Saldos']
        ws.delete_rows(1, ws.max_row)

        ws['A1'] = 'Receita'
        ws['B1'] = '=SUM(\'Recibos de Venda\'!I:I)'
        ws['E1'] = '=B1 - B2'
        ws['A2'] = 'Despesa'
        ws['B2'] = '=SUM(Despesas!F:F)'
        ws['B3'] = 'Débitos'
        ws['C3'] = 'Créditos'
        ws['A4'] = 'LC (Passivo)'
        ws['B4'] = '=SUMIF(\'Lançamentos Contábeis\'!C:C,\"Outros Passivos:Adibras - Transitória de passivos\",\'Lançamentos Contábeis\'!D:D)'
        ws['C4'] = '=SUMIF(\'Lançamentos Contábeis\'!C:C,\"Outros Passivos:Adibras - Transitória de passivos\",\'Lançamentos Contábeis\'!E:E)'
        ws['D4'] = '=C4-B4'
        ws['A5'] = 'Transf e pag'
        ws['B5'] = f'={transferencia}'
        ws['C5'] = f'={pagamentos}'
        ws['A6'] = 'Saldo Planilha'
        ws['B6'] = '=B1-B2+D4-B5'
        ws['A8'] = 'Extratos adibras'
        ws['A9'] = 'Si'
        ws['B9'] = saldo_inicial
        ws['A10'] = 'Sf'
        ws['B10'] = saldo_final
        ws['A11'] = 'Dif'
        ws['B11'] = '=+B10-B9'
        ws['C11'] = '=+C10-C9'
        ws['A12'] = 'Saldo Extratos'
        ws['B12'] = '=+B11+C11'
        ws['A14'] = 'Diferença'
        ws['B14'] = '=B12-B6'

        wb.save(output_path)

        return output_path

    except Exception as e:
        raise ProcessingError(f"Erro durante o processamento: {str(e)}")
